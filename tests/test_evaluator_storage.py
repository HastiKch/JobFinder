"""Tests for evaluator result storage."""

from __future__ import annotations

from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from jobfinder.evaluator.models import JobEvaluation
from jobfinder.evaluator.parsing import ensure_output_columns
from jobfinder.evaluator.storage import (
    row_numbers_to_remove_after_evaluation,
    write_excel_output,
    write_google_output,
)


class FakeRequest:
    """Minimal executable request object for fake Google Sheets calls."""

    def __init__(self, result: dict[str, Any]) -> None:
        self.result = result

    def execute(self) -> dict[str, Any]:
        return self.result


class FakeValuesResource:
    """Fake Google Sheets values resource that records writes."""

    def __init__(self, service: FakeGoogleService) -> None:
        self.service = service

    def get(self, *, spreadsheetId: str, range: str) -> FakeRequest:
        self.service.value_gets.append((spreadsheetId, range))
        return FakeRequest({"values": self.service.values})

    def batchUpdate(
        self,
        *,
        spreadsheetId: str,
        body: dict[str, Any],
    ) -> FakeRequest:
        self.service.value_updates.append((spreadsheetId, body))
        return FakeRequest({})


class FakeSpreadsheetsResource:
    """Fake Google Sheets spreadsheet resource that records dimension updates."""

    def __init__(self, service: FakeGoogleService) -> None:
        self.service = service

    def values(self) -> FakeValuesResource:
        return FakeValuesResource(self.service)

    def get(self, *, spreadsheetId: str, fields: str) -> FakeRequest:
        self.service.metadata_gets.append((spreadsheetId, fields))
        return FakeRequest(
            {"sheets": [{"properties": {"title": "Run", "sheetId": 123}}]}
        )

    def batchUpdate(
        self,
        *,
        spreadsheetId: str,
        body: dict[str, Any],
    ) -> FakeRequest:
        self.service.dimension_updates.append((spreadsheetId, body))
        return FakeRequest({})


class FakeGoogleService:
    """Small fake for the Google Sheets service surface used by storage."""

    def __init__(self, values: list[list[Any]]) -> None:
        self.values = values
        self.value_gets: list[tuple[str, str]] = []
        self.value_updates: list[tuple[str, dict[str, Any]]] = []
        self.metadata_gets: list[tuple[str, str]] = []
        self.dimension_updates: list[tuple[str, dict[str, Any]]] = []

    def spreadsheets(self) -> FakeSpreadsheetsResource:
        return FakeSpreadsheetsResource(self)


def test_write_excel_output_can_skip_cleanup_for_incremental_save(tmp_path):
    """Incremental saves should preserve source columns until final cleanup."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    output_file = tmp_path / "jobs.xlsx"
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])

    worksheet.cell(row=2, column=1).value = "GIS Analyst"
    worksheet.cell(row=2, column=2).value = "Analyze spatial data"

    write_excel_output(
        workbook,
        worksheet,
        output_file,
        headers,
        header_map,
        {
            2: JobEvaluation(
                row_number=2,
                verdict="Suitable",
                fit_score=90,
                reason="Strong match.",
                tailored_cv="CV",
                model="test-model",
            )
        },
        cleanup_columns=False,
    )

    saved = openpyxl.load_workbook(output_file)
    saved_worksheet = saved.active
    assert saved_worksheet.cell(row=1, column=2).value == "Job Description"
    assert saved_worksheet.cell(row=2, column=3).value == "Suitable"

    write_excel_output(
        saved,
        saved_worksheet,
        output_file,
        headers,
        header_map,
        {},
        cleanup_columns=True,
    )

    finalized = openpyxl.load_workbook(output_file)
    finalized_headers = [
        finalized.active.cell(row=1, column=idx).value
        for idx in range(1, finalized.active.max_column + 1)
    ]
    assert "Job Description" not in finalized_headers
    assert finalized.active.cell(row=2, column=2).value == "Suitable"


def test_row_numbers_to_remove_after_evaluation_counts_rejection_labels():
    """Only not-suitable rows with exactly one label should remain."""
    headers = ["Job Title", "AI Verdict", "AI Unsuitable Reasons"]
    rows = [
        ["Keep suitable", "Suitable", ""],
        ["Keep one label", "Not Suitable", "Language proficiency mismatch"],
        [
            "Remove semicolon labels",
            "Not Suitable",
            "Language proficiency mismatch; Degree mismatch",
        ],
        [
            "Remove line labels",
            "Not Suitable",
            "1. Seniority mismatch\n2. Role type mismatch",
        ],
        ["Keep old sentence", "Not Suitable", "Requires fluent German."],
        ["Remove missing label", "Not Suitable", ""],
    ]

    assert row_numbers_to_remove_after_evaluation(headers, rows) == [4, 5, 7]


def test_final_excel_cleanup_removes_multi_label_not_suitable_rows(tmp_path):
    """Final cleanup should hide strongly rejected rows from the saved workbook."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    output_file = tmp_path / "jobs.xlsx"
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])

    worksheet.cell(row=2, column=1).value = "GIS Analyst"
    worksheet.cell(row=2, column=2).value = "Analyze spatial data"
    worksheet.cell(row=3, column=1).value = "German GIS Analyst"
    worksheet.cell(row=3, column=2).value = "German language required"
    worksheet.cell(row=4, column=1).value = "Senior Manager"
    worksheet.cell(row=4, column=2).value = "Senior leadership role"

    write_excel_output(
        workbook,
        worksheet,
        output_file,
        headers,
        header_map,
        {
            2: JobEvaluation(
                row_number=2,
                verdict="Suitable",
                fit_score=90,
                reason="Strong match.",
                tailored_cv="CV",
                model="test-model",
            ),
            3: JobEvaluation(
                row_number=3,
                verdict="Not Suitable",
                fit_score=40,
                reason="German required.",
                unsuitable_reasons="Language proficiency mismatch",
                model="test-model",
            ),
            4: JobEvaluation(
                row_number=4,
                verdict="Not Suitable",
                fit_score=20,
                reason="Too senior and management focused.",
                unsuitable_reasons=(
                    "Seniority mismatch; Leadership/management mismatch"
                ),
                model="test-model",
            ),
        },
        cleanup_columns=True,
    )

    finalized = openpyxl.load_workbook(output_file)
    worksheet = finalized.active
    titles = [
        worksheet.cell(row=row_idx, column=1).value
        for row_idx in range(2, worksheet.max_row + 1)
    ]
    finalized_headers = [
        worksheet.cell(row=1, column=idx).value
        for idx in range(1, worksheet.max_column + 1)
    ]

    assert titles == ["GIS Analyst", "German GIS Analyst"]
    assert "Job Description" not in finalized_headers
    assert worksheet.cell(row=3, column=2).value == "Not Suitable"
    assert worksheet.cell(row=3, column=4).value == "Language proficiency mismatch"


def test_final_excel_cleanup_can_keep_all_not_suitable_rows(tmp_path):
    """The workflow option can preserve every evaluated row."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    output_file = tmp_path / "jobs.xlsx"
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])

    worksheet.cell(row=2, column=1).value = "Senior Manager"
    worksheet.cell(row=2, column=2).value = "Senior leadership role"

    write_excel_output(
        workbook,
        worksheet,
        output_file,
        headers,
        header_map,
        {
            2: JobEvaluation(
                row_number=2,
                verdict="Not Suitable",
                fit_score=20,
                reason="Too senior and management focused.",
                unsuitable_reasons=(
                    "Seniority mismatch; Leadership/management mismatch"
                ),
                model="test-model",
            ),
        },
        cleanup_columns=True,
        remove_rejected_rows=False,
    )

    finalized = openpyxl.load_workbook(output_file)
    worksheet = finalized.active

    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=1).value == "Senior Manager"
    assert worksheet.cell(row=2, column=2).value == "Not Suitable"


def test_final_excel_cleanup_can_remove_tailored_cv_column(tmp_path):
    """Final PDF-enabled cleanup should remove raw LaTeX CV content."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    output_file = tmp_path / "jobs.xlsx"
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])

    worksheet.cell(row=2, column=1).value = "GIS Analyst"
    worksheet.cell(row=2, column=2).value = "Analyze spatial data"

    write_excel_output(
        workbook,
        worksheet,
        output_file,
        headers,
        header_map,
        {
            2: JobEvaluation(
                row_number=2,
                verdict="Suitable",
                fit_score=90,
                reason="Strong match.",
                tailored_cv=r"\documentclass{article}\begin{document}\end{document}",
                cv_pdf="https://drive.google.com/file/d/pdf-id/view",
                model="test-model",
            )
        },
        cleanup_columns=True,
        remove_tailored_cv=True,
    )

    finalized = openpyxl.load_workbook(output_file)
    finalized_headers = [
        finalized.active.cell(row=1, column=idx).value
        for idx in range(1, finalized.active.max_column + 1)
    ]

    assert "AI Tailored CV" not in finalized_headers
    assert "AI CV PDF" in finalized_headers


def test_final_google_cleanup_removes_multi_label_not_suitable_rows():
    """Google Sheets cleanup should delete rejected rows before deleting columns."""
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])
    service = FakeGoogleService(
        [
            headers,
            [
                "Keep one label",
                "Description",
                "Not Suitable",
                40,
                "Language proficiency mismatch",
                "",
            ],
            [
                "Remove two labels",
                "Description",
                "Not Suitable",
                20,
                "Language proficiency mismatch; Degree mismatch",
                "",
            ],
            ["Keep suitable", "Description", "Suitable", 90, "", "CV"],
        ]
    )

    write_google_output(
        service,
        "spreadsheet-id",
        "Run",
        headers,
        header_map,
        {},
        cleanup_columns=True,
    )

    row_update = service.dimension_updates[0][1]
    row_range = row_update["requests"][0]["deleteDimension"]["range"]
    column_update = service.dimension_updates[1][1]
    column_range = column_update["requests"][0]["deleteDimension"]["range"]

    assert row_range == {
        "sheetId": 123,
        "dimension": "ROWS",
        "startIndex": 2,
        "endIndex": 3,
    }
    assert column_range["dimension"] == "COLUMNS"


def test_write_google_output_updates_cv_pdf_column():
    """The final PDF column should receive the generated Drive link."""
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])
    service = FakeGoogleService([headers])
    drive_link = "https://drive.google.com/file/d/pdf-id/view"
    pdf_column = headers.index("AI CV PDF") + 1
    pdf_letter = get_column_letter(pdf_column)

    write_google_output(
        service,
        "spreadsheet-id",
        "Run",
        headers,
        header_map,
        {
            2: JobEvaluation(
                row_number=2,
                verdict="Suitable",
                fit_score=90,
                reason="Strong match.",
                tailored_cv=r"\documentclass{article}\begin{document}\end{document}",
                cv_pdf=drive_link,
                model="test-model",
            )
        },
        cleanup_columns=False,
    )

    data = service.value_updates[0][1]["data"]
    assert {
        "range": f"'Run'!{pdf_letter}2",
        "values": [[drive_link]],
    } in data


def test_final_google_cleanup_can_keep_all_not_suitable_rows():
    """Google Sheets cleanup should skip row deletion when keep_all is selected."""
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])
    service = FakeGoogleService(
        [
            headers,
            [
                "Remove two labels",
                "Description",
                "Not Suitable",
                20,
                "Language proficiency mismatch; Degree mismatch",
                "",
            ],
        ]
    )

    write_google_output(
        service,
        "spreadsheet-id",
        "Run",
        headers,
        header_map,
        {},
        cleanup_columns=True,
        remove_rejected_rows=False,
    )

    assert len(service.dimension_updates) == 1
    column_update = service.dimension_updates[0][1]
    column_range = column_update["requests"][0]["deleteDimension"]["range"]
    assert column_range["dimension"] == "COLUMNS"


def test_final_google_cleanup_can_remove_tailored_cv_column():
    """Final PDF-enabled cleanup should delete the raw LaTeX column in Sheets."""
    headers, header_map = ensure_output_columns(["Job Title", "Job Description"])
    service = FakeGoogleService([headers])

    write_google_output(
        service,
        "spreadsheet-id",
        "Run",
        headers,
        header_map,
        {},
        cleanup_columns=True,
        remove_tailored_cv=True,
    )

    column_update = service.dimension_updates[0][1]
    deleted_indexes = [
        request["deleteDimension"]["range"]["startIndex"]
        for request in column_update["requests"]
    ]
    assert headers.index("AI Tailored CV") in deleted_indexes
    assert headers.index("AI CV PDF") not in deleted_indexes
