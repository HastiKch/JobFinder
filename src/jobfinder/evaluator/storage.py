"""Excel and Google Sheets storage adapters for evaluator results."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from jobfinder.env import EnvSettings
from jobfinder.evaluator.models import (
    DETAIL_COLUMNS,
    OUTPUT_COLUMNS,
    REMOVED_AI_OUTPUT_COLUMNS,
    EvaluationError,
    GoogleSheetsError,
    JobEvaluation,
)
from jobfinder.evaluator.parsing import normalize_header, trim_trailing_blank_headers
from jobfinder.integrations.google.client import google_execute
from jobfinder.integrations.google.sheets import (
    build_google_sheets_service,
    quote_sheet_name,
)
from jobfinder.paths import GOOGLE_SPREADSHEET_ID_FILE

LABEL_SEPARATOR_RE = re.compile(r"[;\n]+")
LIST_MARKER_RE = re.compile(r"^\s*(?:[-*]|\d+[.)])\s*")
GOOGLE_BATCH_REQUEST_CHUNK_SIZE = 500


def resolve_sheet_name(existing_names: list[str], requested: str) -> str:
    """Resolve a requested sheet name, with ``latest`` selecting the newest tab."""
    if not existing_names:
        raise EvaluationError("The workbook/spreadsheet has no sheets.")
    if not requested or requested == "latest":
        return existing_names[-1]
    if requested in existing_names:
        return requested
    raise EvaluationError(
        f"Sheet '{requested}' was not found. Available sheets: "
        f"{', '.join(existing_names)}"
    )


def read_excel_input(
    path: Path,
    requested_sheet: str,
) -> tuple[Any, Any, str, list[str], list[list[Any]]]:
    """Read headers and rows from an Excel workbook sheet."""
    if not path.exists():
        raise EvaluationError(f"Excel file not found: {path}")

    workbook = openpyxl.load_workbook(path)
    sheet_name = resolve_sheet_name(workbook.sheetnames, requested_sheet)
    worksheet = workbook[sheet_name]
    headers = trim_trailing_blank_headers(
        [
            worksheet.cell(row=1, column=col_idx).value
            for col_idx in range(1, worksheet.max_column + 1)
        ]
    )
    if not headers:
        raise EvaluationError(f"Sheet '{sheet_name}' has no header row.")

    rows = [
        [
            worksheet.cell(row=row_idx, column=col_idx).value
            for col_idx in range(1, len(headers) + 1)
        ]
        for row_idx in range(2, worksheet.max_row + 1)
    ]
    return workbook, worksheet, sheet_name, headers, rows


def columns_to_remove_after_evaluation(
    headers: list[str],
    *,
    remove_tailored_cv: bool = False,
) -> list[int]:
    """Return zero-based column indexes to delete after evaluator output is saved."""
    removable_columns = [*REMOVED_AI_OUTPUT_COLUMNS, *DETAIL_COLUMNS]
    if remove_tailored_cv:
        removable_columns.append("AI Tailored CV")

    removable = {normalize_header(column) for column in removable_columns}
    return [
        idx
        for idx, header in enumerate(headers)
        if normalize_header(header) in removable
    ]


def remove_excel_columns_after_evaluation(
    worksheet: Any,
    headers: list[str],
    *,
    remove_tailored_cv: bool = False,
) -> None:
    """Delete legacy AI metadata and job detail columns from an Excel worksheet."""
    for column_idx in sorted(
        columns_to_remove_after_evaluation(
            headers,
            remove_tailored_cv=remove_tailored_cv,
        ),
        reverse=True,
    ):
        worksheet.delete_cols(column_idx + 1)


def header_index(headers: list[str], column_name: str) -> int | None:
    """Return the zero-based index for a normalized header name."""
    normalized_name = normalize_header(column_name)
    for idx, header in enumerate(headers):
        if normalize_header(header) == normalized_name:
            return idx
    return None


def cell_text(value: Any) -> str:
    """Return a normalized text value for row-removal decisions."""
    return "" if value is None else str(value).strip()


def unsuitable_reason_label_count(value: Any) -> int:
    """Count semicolon or line separated unsuitable-reason labels."""
    text = cell_text(value)
    if not text:
        return 0
    labels = []
    for part in LABEL_SEPARATOR_RE.split(text):
        label = LIST_MARKER_RE.sub("", part).strip()
        if label:
            labels.append(label)
    return len(labels)


def should_remove_after_evaluation(verdict: Any, unsuitable_reasons: Any) -> bool:
    """Return true for not-suitable rows without exactly one rejection label."""
    if cell_text(verdict).casefold() != "not suitable":
        return False
    return unsuitable_reason_label_count(unsuitable_reasons) != 1


def row_value(row: list[Any], idx: int | None) -> Any:
    """Return a row value or an empty string when the index is unavailable."""
    if idx is None or idx >= len(row):
        return ""
    return row[idx]


def row_numbers_to_remove_after_evaluation(
    headers: list[str],
    rows: list[list[Any]],
) -> list[int]:
    """Return one-based worksheet row numbers to remove after final evaluation."""
    verdict_idx = header_index(headers, "AI Verdict")
    unsuitable_reasons_idx = header_index(headers, "AI Unsuitable Reasons")
    if verdict_idx is None or unsuitable_reasons_idx is None:
        return []

    return [
        row_number
        for row_number, row in enumerate(rows, start=2)
        if should_remove_after_evaluation(
            row_value(row, verdict_idx),
            row_value(row, unsuitable_reasons_idx),
        )
    ]


def remove_excel_rows_after_evaluation(worksheet: Any, headers: list[str]) -> None:
    """Delete not-suitable Excel rows unless they have exactly one label."""
    verdict_idx = header_index(headers, "AI Verdict")
    unsuitable_reasons_idx = header_index(headers, "AI Unsuitable Reasons")
    if verdict_idx is None or unsuitable_reasons_idx is None:
        return

    for row_idx in range(worksheet.max_row, 1, -1):
        if should_remove_after_evaluation(
            worksheet.cell(row=row_idx, column=verdict_idx + 1).value,
            worksheet.cell(row=row_idx, column=unsuitable_reasons_idx + 1).value,
        ):
            worksheet.delete_rows(row_idx)


def write_excel_output(
    workbook: Any,
    worksheet: Any,
    path: Path,
    headers: list[str],
    header_map: dict[str, int],
    evaluations: dict[int, JobEvaluation],
    *,
    cleanup_columns: bool = True,
    remove_rejected_rows: bool = True,
    remove_tailored_cv: bool = False,
) -> None:
    """Write evaluator columns and results back to an Excel worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx).value = header

    for evaluation in evaluations.values():
        for column in OUTPUT_COLUMNS:
            column_idx = header_map[normalize_header(column)] + 1
            worksheet.cell(
                row=evaluation.row_number,
                column=column_idx,
            ).value = evaluation.value_for_column(column)

    if cleanup_columns:
        if remove_rejected_rows:
            remove_excel_rows_after_evaluation(worksheet, headers)
        remove_excel_columns_after_evaluation(
            worksheet,
            headers,
            remove_tailored_cv=remove_tailored_cv,
        )
    workbook.save(path)


def build_evaluator_google_sheets_service() -> Any:
    """Build a Google Sheets service for evaluator reads and writes."""
    return build_google_sheets_service(error_cls=GoogleSheetsError)


def read_google_spreadsheet_id(cli_value: str) -> str:
    """Resolve a spreadsheet ID from CLI, env, or local cache file."""
    if cli_value:
        return cli_value

    env_value = EnvSettings().get("GOOGLE_SPREADSHEET_ID")
    if env_value:
        return env_value
    if GOOGLE_SPREADSHEET_ID_FILE.exists():
        return GOOGLE_SPREADSHEET_ID_FILE.read_text(encoding="utf-8").strip()
    return ""


def read_google_input(
    service: Any,
    spreadsheet_id: str,
    requested_sheet: str,
) -> tuple[str, list[str], list[list[Any]]]:
    """Read headers and rows from a Google Sheet tab."""
    try:
        metadata = google_execute(
            service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                fields="sheets(properties(title))",
            )
        )
        sheet_names = [
            sheet["properties"]["title"] for sheet in metadata.get("sheets", [])
        ]
        sheet_name = resolve_sheet_name(sheet_names, requested_sheet)
        response = google_execute(
            service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=quote_sheet_name(sheet_name))
        )
    except Exception as exc:
        raise GoogleSheetsError(
            f"Could not read Google spreadsheet ID '{spreadsheet_id}'. Details: {exc}"
        ) from exc

    values = response.get("values", [])
    if not values:
        raise EvaluationError(f"Google Sheet tab '{sheet_name}' is empty.")

    headers = trim_trailing_blank_headers(values[0])
    if not headers:
        raise EvaluationError(f"Google Sheet tab '{sheet_name}' has no header row.")
    rows = [list(row) for row in values[1:]]
    return sheet_name, headers, rows


def write_google_output(
    service: Any,
    spreadsheet_id: str,
    sheet_name: str,
    headers: list[str],
    header_map: dict[str, int],
    evaluations: dict[int, JobEvaluation],
    *,
    cleanup_columns: bool = True,
    remove_rejected_rows: bool = True,
    remove_tailored_cv: bool = False,
) -> None:
    """Write evaluator columns and results back to a Google Sheet tab."""
    try:
        data = []
        for column in OUTPUT_COLUMNS:
            column_idx = header_map[normalize_header(column)]
            column_letter = get_column_letter(column_idx + 1)
            data.append(
                {
                    "range": f"{quote_sheet_name(sheet_name)}!{column_letter}1",
                    "values": [[headers[column_idx]]],
                }
            )
            data.extend(
                {
                    "range": (
                        f"{quote_sheet_name(sheet_name)}!"
                        f"{column_letter}{evaluation.row_number}"
                    ),
                    "values": [[evaluation.value_for_column(column)]],
                }
                for evaluation in evaluations.values()
            )

        for idx in range(0, len(data), GOOGLE_BATCH_REQUEST_CHUNK_SIZE):
            google_execute(
                service.spreadsheets()
                .values()
                .batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        "valueInputOption": "RAW",
                        "data": data[idx : idx + GOOGLE_BATCH_REQUEST_CHUNK_SIZE],
                    },
                )
            )

        if cleanup_columns:
            if remove_rejected_rows:
                remove_google_rows_after_evaluation(
                    service,
                    spreadsheet_id,
                    sheet_name,
                )
            remove_google_columns_after_evaluation(
                service,
                spreadsheet_id,
                sheet_name,
                headers,
                remove_tailored_cv=remove_tailored_cv,
            )
    except Exception as exc:
        raise GoogleSheetsError(
            f"Could not write Google spreadsheet ID '{spreadsheet_id}'. Details: {exc}"
        ) from exc


def get_google_sheet_id(service: Any, spreadsheet_id: str, sheet_name: str) -> int:
    """Return the numeric Google Sheets tab ID for a sheet title."""
    metadata = google_execute(
        service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title))",
        )
    )
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if properties.get("title") == sheet_name:
            return int(properties["sheetId"])
    raise EvaluationError(f"Google Sheet tab '{sheet_name}' was not found.")


def remove_google_rows_after_evaluation(
    service: Any,
    spreadsheet_id: str,
    sheet_name: str,
) -> None:
    """Delete not-suitable Google Sheet rows unless they have exactly one label."""
    response = google_execute(
        service.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=quote_sheet_name(sheet_name))
    )
    values = response.get("values", [])
    if not values:
        return

    headers = trim_trailing_blank_headers(values[0])
    rows = [list(row) for row in values[1:]]
    row_numbers = row_numbers_to_remove_after_evaluation(headers, rows)
    if not row_numbers:
        return

    sheet_id = get_google_sheet_id(service, spreadsheet_id, sheet_name)
    requests = [
        {
            "deleteDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": row_number - 1,
                    "endIndex": row_number,
                }
            }
        }
        for row_number in sorted(row_numbers, reverse=True)
    ]
    batch_update_google_requests(service, spreadsheet_id, requests)


def remove_google_columns_after_evaluation(
    service: Any,
    spreadsheet_id: str,
    sheet_name: str,
    headers: list[str],
    *,
    remove_tailored_cv: bool = False,
) -> None:
    """Delete legacy AI metadata and job detail columns from a Google Sheet tab."""
    column_indexes = columns_to_remove_after_evaluation(
        headers,
        remove_tailored_cv=remove_tailored_cv,
    )
    if not column_indexes:
        return

    sheet_id = get_google_sheet_id(service, spreadsheet_id, sheet_name)
    requests = [
        {
            "deleteDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": column_idx,
                    "endIndex": column_idx + 1,
                }
            }
        }
        for column_idx in sorted(column_indexes, reverse=True)
    ]
    batch_update_google_requests(service, spreadsheet_id, requests)


def batch_update_google_requests(
    service: Any,
    spreadsheet_id: str,
    requests: list[dict[str, Any]],
) -> None:
    """Send Google Sheets batchUpdate requests in bounded chunks."""
    for idx in range(0, len(requests), GOOGLE_BATCH_REQUEST_CHUNK_SIZE):
        google_execute(
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body={
                    "requests": requests[idx : idx + GOOGLE_BATCH_REQUEST_CHUNK_SIZE]
                },
            ),
            retries=0,
        )
