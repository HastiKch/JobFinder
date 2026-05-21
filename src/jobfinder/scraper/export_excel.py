"""Excel export support for scraper results."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from jobfinder.scraper.export_rows import HEADER, make_job_rows, unique_name
from jobfinder.scraper.settings import ScraperSettings

COLOR_HEADER_BG = "102C53"
COLOR_HEADER_FG = "FFFFFF"
COLOR_ROW_ODD = "CADCFC"
COLOR_ROW_EVEN = "FFFFFF"
COLOR_ACCENT = "C9A84C"

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_cell(cell: Any) -> None:
    """Apply the standard header style to an Excel cell."""
    cell.font = Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_data_cell(cell: Any, row_idx: int, is_url: bool = False) -> None:
    """Apply the standard row style to an Excel cell."""
    bg = COLOR_ROW_ODD if row_idx % 2 == 1 else COLOR_ROW_EVEN
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = BORDER
    if is_url:
        cell.font = Font(
            color=COLOR_ACCENT, name="Calibri", size=10, underline="single"
        )
    else:
        cell.font = Font(name="Calibri", size=10)


def parse_hyperlink_formula(value: str) -> tuple[str, str] | None:
    """Parse a spreadsheet hyperlink formula into URL and label parts."""
    prefix = '=HYPERLINK("'
    separator = '", "'
    suffix = '")'
    if (
        not isinstance(value, str)
        or not value.startswith(prefix)
        or not value.endswith(suffix)
    ):
        return None

    body = value[len(prefix) : -len(suffix)]
    if separator not in body:
        return None

    url, label = body.split(separator, 1)
    return url.replace('""', '"'), label.replace('""', '"')


def excel_value(value: Any) -> Any:
    """Return the display value to write into an Excel cell."""
    parsed = parse_hyperlink_formula(value)
    if parsed:
        return parsed[1]
    return value


def export_to_excel(
    settings: ScraperSettings, jobs: list[dict[str, Any]], filename: Path
) -> str:
    """Write jobs to a new timestamped worksheet in an Excel workbook."""
    if filename.exists():
        wb = openpyxl.load_workbook(filename)
        sheet_name = unique_name(set(wb.sheetnames), settings.run_sheet_name, 31)
        ws = wb.create_sheet(sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = unique_name(set(), settings.run_sheet_name, 31)

    job_rows = make_job_rows(settings, jobs)
    for row in job_rows:
        ws.append([excel_value(value) for value in row])

    for cell in ws[1]:
        style_header_cell(cell)
    ws.row_dimensions[1].height = 30

    url_columns = {idx for idx, name in enumerate(HEADER, start=1) if "URL" in name}
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            style_data_cell(cell, row_idx, is_url=col_idx in url_columns)
            parsed = parse_hyperlink_formula(job_rows[row_idx - 1][col_idx - 1])
            if parsed:
                cell.hyperlink = parsed[0]

    width_by_header = {
        "Application Status": 20,
        "App": 14,
        "Job Title": 40,
        "Company": 32,
        "Location": 28,
        "Job Type": 18,
        "Job Description": 70,
        "Posted": 22,
        "Keywords Matched": 32,
        "Job URL": 18,
        "Apply URL": 18,
        "AI Verdict": 18,
        "AI Fit Score": 14,
        "AI Unsuitable Reasons": 60,
        "AI Tailored CV": 60,
        "AI CV PDF": 60,
    }
    for col_idx, header in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_by_header.get(
            header, 18
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    return f"{filename} (sheet: {ws.title})"
